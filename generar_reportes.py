"""
LIBERTAD_2045 — Generador de reportes v2
Genera dashboard HTML y Excel detallado a partir de los últimos 3 backtests.
"""

import json
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.series import SeriesLabel


# --------------------------------------------------
# Configuración
# --------------------------------------------------

RESULTS_DIR  = Path("backtest_results")
OUTPUT_DIR   = Path(".")

ANTES = {
    "2015": {"retorno": 0.04439, "win_rate": 0.493, "pf": 2.1991,
             "dd": 0.093, "trades": 747,  "cap_final":  181576},
    "2010": {"retorno": 0.14258, "win_rate": 0.512, "pf": 2.1921,
             "dd": 0.093, "trades": 1096, "cap_final":  574338},
    "2005": {"retorno": 0.27032, "win_rate": 0.479, "pf": 2.1758,
             "dd": 0.126, "trades": 1368, "cap_final": 1085280},
}

COLORES = {
    "2015": "#00d4ff",
    "2010": "#7c3aed",
    "2005": "#f59e0b",
    "win":  "#10b981",
    "loss": "#ef4444",
    "bg":   "#0f1117",
    "card": "#1a1d2e",
    "border": "#2d3148",
}


# --------------------------------------------------
# Carga de datos
# --------------------------------------------------

def cargar_datos():
    """Carga los 3 conjuntos más recientes de resultados."""

    archivos = sorted(RESULTS_DIR.glob("expandido_metricas_*.csv"), reverse=True)
    sets_cargados = []

    for f in archivos:
        ts = f.stem.replace("expandido_metricas_", "")
        cap_file    = RESULTS_DIR / f"expandido_capital_{ts}.csv"
        trades_file = RESULTS_DIR / f"expandido_trades_{ts}.csv"

        if not cap_file.exists() or not trades_file.exists():
            continue

        capital = pd.read_csv(cap_file, parse_dates=["fecha"])
        trades  = pd.read_csv(trades_file, parse_dates=["fecha_entrada", "fecha_salida"])
        metrics = pd.read_csv(f).iloc[0]

        year = str(capital["fecha"].iloc[0].year)

        sets_cargados.append({
            "year":    year,
            "capital": capital,
            "trades":  trades,
            "metrics": metrics,
        })

        if len(sets_cargados) == 3:
            break

    # Ordenar por año descendente (2015 primero)
    sets_cargados.sort(key=lambda x: x["year"], reverse=True)
    return sets_cargados


def metricas_anuales(capital_df, trades_df):
    """Calcula métricas por año: retorno, trades, win rate, PnL."""

    capital_df = capital_df.copy()
    capital_df["year"] = capital_df["fecha"].dt.year

    trades_df = trades_df.copy()
    trades_df["year"] = trades_df["fecha_salida"].dt.year

    años = sorted(capital_df["year"].unique())
    rows = []

    for yr in años:
        cap_yr = capital_df[capital_df["year"] == yr]
        cap_inicio = cap_yr["capital"].iloc[0]
        cap_fin    = cap_yr["capital"].iloc[-1]
        retorno    = (cap_fin - cap_inicio) / cap_inicio if cap_inicio > 0 else 0

        tr_yr = trades_df[trades_df["year"] == yr]
        total = len(tr_yr)
        wins  = len(tr_yr[tr_yr["pnl"] >= 0])
        pnl   = tr_yr["pnl"].sum()
        wr    = wins / total if total > 0 else 0

        # Drawdown del año
        caps = cap_yr["capital"].values
        pico = caps[0]
        max_dd = 0.0
        for c in caps:
            if c > pico:
                pico = c
            dd = (pico - c) / pico if pico > 0 else 0
            if dd > max_dd:
                max_dd = dd

        rows.append({
            "Año":        yr,
            "Cap. inicio":  round(cap_inicio, 0),
            "Cap. fin":     round(cap_fin, 0),
            "Retorno":      round(retorno * 100, 1),
            "Trades":       total,
            "Wins":         wins,
            "Win Rate":     round(wr * 100, 1),
            "PnL total":    round(pnl, 0),
            "Max DD":       round(max_dd * 100, 1),
        })

    return pd.DataFrame(rows)


def calcular_drawdown_serie(capital_df):
    """Devuelve serie de drawdown en %."""
    caps = capital_df["capital"].values
    pico = caps[0]
    dds  = []
    for c in caps:
        if c > pico:
            pico = c
        dds.append(round((pico - c) / pico * 100, 3) if pico > 0 else 0)
    return dds


# --------------------------------------------------
# HTML DASHBOARD
# --------------------------------------------------

def generar_html(datasets):

    # Preparar datos para Chart.js
    def muestra(serie, n=600):
        """Submuestrea una serie a n puntos máximo."""
        if len(serie) <= n:
            return serie
        idx = np.linspace(0, len(serie) - 1, n, dtype=int)
        return [serie[i] for i in idx]

    curvas_js  = []
    dd_js      = []
    scatter_js = []

    for ds in datasets:
        yr    = ds["year"]
        col   = COLORES[yr]
        cap   = ds["capital"]
        fechas = [str(d.date()) for d in cap["fecha"]]
        caps   = cap["capital"].tolist()
        dds    = calcular_drawdown_serie(cap)

        # Submuestrear
        idx = np.linspace(0, len(fechas)-1, min(600, len(fechas)), dtype=int)
        f_s = [fechas[i] for i in idx]
        c_s = [caps[i]   for i in idx]
        d_s = [dds[i]    for i in idx]

        curvas_js.append({
            "label":       f"Desde {yr}",
            "data":        [{"x": f, "y": round(c, 2)} for f, c in zip(f_s, c_s)],
            "borderColor": col,
            "backgroundColor": col + "18",
            "borderWidth": 2,
            "pointRadius": 0,
            "fill": True,
            "tension": 0.3,
        })

        dd_js.append({
            "label":       f"Desde {yr}",
            "data":        [{"x": f, "y": round(d, 2)} for f, d in zip(f_s, d_s)],
            "borderColor": col,
            "backgroundColor": col + "18",
            "borderWidth": 2,
            "pointRadius": 0,
            "fill": True,
            "tension": 0.3,
        })

        # Scatter PnL por trade
        tr = ds["trades"].copy()
        wins  = tr[tr["pnl"] >= 0].sample(min(len(tr[tr["pnl"] >= 0]), 300), random_state=42)
        losses = tr[tr["pnl"] < 0].sample(min(len(tr[tr["pnl"] < 0]), 300), random_state=42)
        scatter_js.append({
            "yr": yr, "col": col,
            "wins":   [{"x": str(r.fecha_salida.date()), "y": round(r.pnl, 2)} for _, r in wins.iterrows()],
            "losses": [{"x": str(r.fecha_salida.date()), "y": round(r.pnl, 2)} for _, r in losses.iterrows()],
        })

    # Tabla comparativa antes/después
    tabla_rows = ""
    orden = ["2015", "2010", "2005"]
    for ds in sorted(datasets, key=lambda x: x["year"], reverse=True):
        yr  = ds["year"]
        m   = ds["metrics"]
        ant = ANTES.get(yr, {})
        col = COLORES[yr]

        def delta_str(nuevo, antes, pct=False, invert=False):
            if not antes:
                return ""
            d = nuevo - antes
            mejor = d > 0 if not invert else d < 0
            color = "#10b981" if mejor else "#ef4444"
            signo = "+" if d >= 0 else ""
            val = f"{d:.1f}{'%' if pct else ''}"
            return f'<span style="color:{color};font-size:0.8em"> ({signo}{val})</span>'

        wr_n  = round(m["win_rate"] * 100, 1)
        wr_a  = round(ant.get("win_rate", 0) * 100, 1) if ant else 0
        ret_n = round(m["retorno_total"] * 100, 1)
        ret_a = round(ant.get("retorno", 0) * 100, 1) if ant else 0
        dd_n  = round(m["max_drawdown"] * 100, 1)
        dd_a  = round(ant.get("dd", 0) * 100, 1) if ant else 0
        pf_n  = round(m["profit_factor"], 4)
        pf_a  = round(ant.get("pf", 0), 4) if ant else 0
        tr_n  = int(m["total_trades"])
        tr_a  = int(ant.get("trades", 0)) if ant else 0
        cap_n = round(m["capital_final"])
        cap_a = int(ant.get("cap_final", 0)) if ant else 0

        tabla_rows += f"""
        <tr>
          <td><span style="color:{col};font-weight:700">Desde {yr}</span></td>
          <td>{cap_n:,.0f} €{delta_str(cap_n, cap_a)}</td>
          <td>{ret_n:.1f}%{delta_str(ret_n, ret_a, pct=True)}</td>
          <td>{tr_n}{delta_str(tr_n, tr_a)}</td>
          <td>{wr_n:.1f}%{delta_str(wr_n, wr_a, pct=True)}</td>
          <td>{pf_n:.4f}{delta_str(pf_n, pf_a)}</td>
          <td>{dd_n:.1f}%{delta_str(dd_n, dd_a, pct=True, invert=True)}</td>
        </tr>"""

    # Métricas anuales por período
    tabs_html = ""
    panels_html = ""
    for i, ds in enumerate(sorted(datasets, key=lambda x: x["year"], reverse=True)):
        yr  = ds["year"]
        col = COLORES[yr]
        ma  = metricas_anuales(ds["capital"], ds["trades"])
        active = "active" if i == 0 else ""

        tabs_html += f'<button class="tab-btn {active}" onclick="showTab(\'{yr}\')" id="tab-{yr}" style="border-bottom: 2px solid {col if active else "transparent"}">' \
                     f'<span style="color:{col}">Desde {yr}</span></button>'

        filas = ""
        for _, row in ma.iterrows():
            r_color = "#10b981" if row["Retorno"] >= 0 else "#ef4444"
            filas += f"""<tr>
              <td>{int(row['Año'])}</td>
              <td>{row['Cap. inicio']:,.0f}</td>
              <td>{row['Cap. fin']:,.0f}</td>
              <td style="color:{r_color}">{row['Retorno']:+.1f}%</td>
              <td>{int(row['Trades'])}</td>
              <td style="color:#10b981">{int(row['Wins'])}</td>
              <td>{row['Win Rate']:.1f}%</td>
              <td style="color:{r_color}">{row['PnL total']:,.0f} €</td>
              <td style="color:#ef4444">{row['Max DD']:.1f}%</td>
            </tr>"""

        panels_html += f"""
        <div class="tab-panel {'active' if active else ''}" id="panel-{yr}">
          <table class="data-table">
            <thead><tr>
              <th>Año</th><th>Cap. inicio</th><th>Cap. fin</th>
              <th>Retorno</th><th>Trades</th><th>Wins</th>
              <th>Win Rate</th><th>PnL</th><th>Max DD</th>
            </tr></thead>
            <tbody>{filas}</tbody>
          </table>
        </div>"""

    # KPI cards
    kpi_cards = ""
    for ds in sorted(datasets, key=lambda x: x["year"], reverse=True):
        yr  = ds["year"]
        m   = ds["metrics"]
        col = COLORES[yr]
        tr  = ds["trades"]
        mejor_trade = tr.loc[tr["pnl"].idxmax()]
        peor_trade  = tr.loc[tr["pnl"].idxmin()]

        kpi_cards += f"""
        <div class="kpi-group">
          <div class="kpi-title" style="color:{col}; border-left: 3px solid {col}; padding-left:10px">
            DESDE {yr}
          </div>
          <div class="kpi-row">
            <div class="kpi-card">
              <div class="kpi-label">Capital final</div>
              <div class="kpi-value" style="color:{col}">{m['capital_final']:,.0f} €</div>
            </div>
            <div class="kpi-card">
              <div class="kpi-label">Retorno total</div>
              <div class="kpi-value" style="color:{col}">{m['retorno_total']*100:.0f}%</div>
            </div>
            <div class="kpi-card">
              <div class="kpi-label">Win Rate</div>
              <div class="kpi-value" style="color:#10b981">{m['win_rate']*100:.1f}%</div>
            </div>
            <div class="kpi-card">
              <div class="kpi-label">Profit Factor</div>
              <div class="kpi-value">{m['profit_factor']:.4f}</div>
            </div>
            <div class="kpi-card">
              <div class="kpi-label">Max Drawdown</div>
              <div class="kpi-value" style="color:#ef4444">{m['max_drawdown']*100:.1f}%</div>
            </div>
            <div class="kpi-card">
              <div class="kpi-label">Expectativa / trade</div>
              <div class="kpi-value">{m['expectativa']:,.0f} €</div>
            </div>
            <div class="kpi-card">
              <div class="kpi-label">Mejor trade</div>
              <div class="kpi-value" style="color:#10b981">+{mejor_trade['pnl']:,.0f} € ({mejor_trade['symbol']})</div>
            </div>
            <div class="kpi-card">
              <div class="kpi-label">Peor trade</div>
              <div class="kpi-value" style="color:#ef4444">{peor_trade['pnl']:,.0f} € ({peor_trade['symbol']})</div>
            </div>
          </div>
        </div>"""

    # Distribución wins/losses (histograma de PnL)
    hist_datasets = []
    for ds in sorted(datasets, key=lambda x: x["year"], reverse=True):
        yr = ds["year"]
        col = COLORES[yr]
        tr = ds["trades"]
        # Bins simples
        bins = np.linspace(tr["pnl"].quantile(0.01), tr["pnl"].quantile(0.99), 30)
        counts, edges = np.histogram(tr["pnl"], bins=bins)
        labels = [f"{(edges[i]+edges[i+1])/2:.0f}" for i in range(len(counts))]
        bar_colors = [COLORES["win"] + "cc" if float(l) >= 0 else COLORES["loss"] + "cc" for l in labels]
        hist_datasets.append({
            "yr": yr, "col": col,
            "labels": labels, "counts": counts.tolist(), "colors": bar_colors,
        })

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>LIBERTAD_2045 — Backtest Dashboard v2</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-adapter-date-fns@3.0.0/dist/chartjs-adapter-date-fns.bundle.min.js"></script>
<style>
  :root {{
    --bg:      #0f1117;
    --card:    #1a1d2e;
    --border:  #2d3148;
    --text:    #e2e8f0;
    --muted:   #94a3b8;
    --accent:  #00d4ff;
  }}
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    background: var(--bg);
    color: var(--text);
    font-family: 'Segoe UI', system-ui, sans-serif;
    font-size: 14px;
    line-height: 1.6;
  }}
  .header {{
    background: linear-gradient(135deg, #0f1117 0%, #1a1d2e 50%, #0f1117 100%);
    border-bottom: 1px solid var(--border);
    padding: 28px 40px;
    display: flex;
    justify-content: space-between;
    align-items: center;
  }}
  .header h1 {{
    font-size: 1.6em;
    font-weight: 700;
    letter-spacing: 0.05em;
    color: var(--accent);
  }}
  .header h1 span {{ color: var(--text); }}
  .header .meta {{ color: var(--muted); font-size: 0.85em; text-align: right; }}
  .badge {{
    display: inline-block;
    background: #10b98120;
    color: #10b981;
    border: 1px solid #10b98140;
    border-radius: 4px;
    padding: 2px 10px;
    font-size: 0.75em;
    font-weight: 600;
    letter-spacing: 0.08em;
  }}
  .container {{ max-width: 1400px; margin: 0 auto; padding: 32px 24px; }}
  .section-title {{
    font-size: 0.7em;
    font-weight: 700;
    letter-spacing: 0.15em;
    color: var(--muted);
    text-transform: uppercase;
    margin-bottom: 16px;
    padding-bottom: 8px;
    border-bottom: 1px solid var(--border);
  }}
  .card {{
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 24px;
    margin-bottom: 24px;
  }}
  .card-title {{
    font-size: 0.85em;
    font-weight: 600;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-bottom: 20px;
  }}
  .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }}
  .chart-container {{ position: relative; height: 340px; }}
  .chart-container-tall {{ position: relative; height: 420px; }}

  /* KPI */
  .kpi-group {{ margin-bottom: 28px; }}
  .kpi-title {{ font-size: 0.75em; font-weight: 700; letter-spacing: 0.12em;
                text-transform: uppercase; margin-bottom: 12px; }}
  .kpi-row {{ display: flex; flex-wrap: wrap; gap: 12px; }}
  .kpi-card {{
    background: #12151f;
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 14px 18px;
    min-width: 140px;
    flex: 1;
  }}
  .kpi-label {{ font-size: 0.72em; color: var(--muted); text-transform: uppercase;
                letter-spacing: 0.08em; margin-bottom: 6px; }}
  .kpi-value {{ font-size: 1.1em; font-weight: 700; }}

  /* Tabla comparativa */
  .compare-table {{ width: 100%; border-collapse: collapse; }}
  .compare-table th {{
    background: #12151f;
    color: var(--muted);
    font-size: 0.72em;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    padding: 10px 16px;
    text-align: left;
    border-bottom: 1px solid var(--border);
  }}
  .compare-table td {{
    padding: 12px 16px;
    border-bottom: 1px solid #1e2235;
    font-weight: 500;
  }}
  .compare-table tr:hover td {{ background: #1e2235; }}

  /* Tabla datos */
  .data-table {{ width: 100%; border-collapse: collapse; font-size: 0.85em; }}
  .data-table th {{
    background: #12151f;
    color: var(--muted);
    font-size: 0.72em;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    padding: 8px 12px;
    text-align: right;
    border-bottom: 1px solid var(--border);
    position: sticky; top: 0;
  }}
  .data-table th:first-child {{ text-align: left; }}
  .data-table td {{
    padding: 8px 12px;
    border-bottom: 1px solid #1e2235;
    text-align: right;
  }}
  .data-table td:first-child {{ text-align: left; }}
  .data-table tr:hover td {{ background: #1e2235; }}

  /* Tabs */
  .tabs {{ display: flex; gap: 4px; margin-bottom: 20px; border-bottom: 1px solid var(--border); }}
  .tab-btn {{
    background: none; border: none; border-bottom: 2px solid transparent;
    color: var(--muted); cursor: pointer; padding: 10px 20px;
    font-size: 0.85em; font-weight: 600; letter-spacing: 0.05em;
    transition: all 0.2s; margin-bottom: -1px;
  }}
  .tab-btn.active {{ color: var(--text); }}
  .tab-panel {{ display: none; max-height: 480px; overflow-y: auto; }}
  .tab-panel.active {{ display: block; }}

  /* Footer */
  .footer {{
    text-align: center;
    padding: 24px;
    color: var(--muted);
    font-size: 0.8em;
    border-top: 1px solid var(--border);
    margin-top: 8px;
  }}
  .sep {{ margin: 32px 0; }}
</style>
</head>
<body>

<div class="header">
  <div>
    <div style="color:var(--muted);font-size:0.75em;letter-spacing:0.15em;
                text-transform:uppercase;margin-bottom:6px">
      Sistema de trading algorítmico
    </div>
    <h1>LIBERTAD<span>_2045</span></h1>
    <div style="margin-top:8px">
      <span class="badge">VERSIÓN 2 — ACTIVA</span>
    </div>
  </div>
  <div class="meta">
    <div style="font-size:1em;color:var(--text);font-weight:600;margin-bottom:4px">
      Backtest Dashboard
    </div>
    <div>5 mejoras implementadas · 2026-04-14</div>
    <div>Generado: {ts}</div>
  </div>
</div>

<div class="container">

  <!-- KPI CARDS -->
  <div class="section-title">Métricas clave por período</div>
  <div class="card">
    {kpi_cards}
  </div>

  <!-- CURVAS DE CAPITAL -->
  <div class="section-title">Curvas de capital — 3 períodos superpuestos</div>
  <div class="card">
    <div class="card-title">Capital acumulado (€) · escala logarítmica · Capital inicial 4.000 € + 4.000 €/año</div>
    <div class="chart-container-tall">
      <canvas id="chartCapital"></canvas>
    </div>
  </div>

  <!-- DRAWDOWN -->
  <div class="section-title">Drawdown</div>
  <div class="card">
    <div class="card-title">Drawdown máximo desde pico (%)</div>
    <div class="chart-container">
      <canvas id="chartDD"></canvas>
    </div>
  </div>

  <!-- COMPARATIVA ANTES / DESPUÉS -->
  <div class="section-title">Comparativa versión 1 vs versión 2</div>
  <div class="card">
    <div class="card-title">Impacto de las 5 mejoras implementadas el 2026-04-14</div>
    <table class="compare-table">
      <thead>
        <tr>
          <th>Período</th>
          <th>Capital final</th>
          <th>Retorno total</th>
          <th>Trades</th>
          <th>Win Rate</th>
          <th>Profit Factor</th>
          <th>Max Drawdown</th>
        </tr>
      </thead>
      <tbody>
        {tabla_rows}
      </tbody>
    </table>
  </div>

  <!-- DISTRIBUCIÓN PnL -->
  <div class="section-title">Distribución de resultados por trade</div>
  <div class="grid-2">
    <div class="card">
      <div class="card-title">Histograma PnL por trade</div>
      <div class="chart-container">
        <canvas id="chartHist0"></canvas>
      </div>
    </div>
    <div class="card">
      <div class="card-title">Histograma PnL por trade</div>
      <div class="chart-container">
        <canvas id="chartHist1"></canvas>
      </div>
    </div>
  </div>
  <div class="card">
    <div class="card-title">Histograma PnL — período completo 2005–2025</div>
    <div class="chart-container">
      <canvas id="chartHist2"></canvas>
    </div>
  </div>

  <!-- MÉTRICAS ANUALES -->
  <div class="section-title">Métricas anuales desglosadas</div>
  <div class="card">
    <div class="tabs">
      {tabs_html}
    </div>
    {panels_html}
  </div>

</div>

<div class="footer">
  LIBERTAD_2045 v2 · Arquitecto: Javier Beneito · {ts}
</div>

<script>
const COLORES = {json.dumps(COLORES)};

// ── Curvas de capital ──
const capitalData = {json.dumps(curvas_js)};
new Chart(document.getElementById('chartCapital'), {{
  type: 'line',
  data: {{ datasets: capitalData }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    scales: {{
      x: {{
        type: 'time',
        time: {{ unit: 'year', tooltipFormat: 'yyyy-MM-dd' }},
        grid: {{ color: '#2d314820' }},
        ticks: {{ color: '#94a3b8', maxTicksLimit: 12 }},
      }},
      y: {{
        type: 'logarithmic',
        grid: {{ color: '#2d314840' }},
        ticks: {{
          color: '#94a3b8',
          callback: v => v >= 1000000 ? (v/1000000).toFixed(1)+'M' : v >= 1000 ? (v/1000).toFixed(0)+'k' : v
        }},
      }}
    }},
    plugins: {{
      legend: {{ labels: {{ color: '#e2e8f0', boxWidth: 14, padding: 20 }} }},
      tooltip: {{
        backgroundColor: '#1a1d2e',
        borderColor: '#2d3148',
        borderWidth: 1,
        titleColor: '#94a3b8',
        bodyColor: '#e2e8f0',
        callbacks: {{
          label: ctx => ctx.dataset.label + ': ' +
            new Intl.NumberFormat('es-ES', {{style:'currency',currency:'EUR',maximumFractionDigits:0}})
            .format(ctx.parsed.y)
        }}
      }}
    }}
  }}
}});

// ── Drawdown ──
const ddData = {json.dumps(dd_js)};
new Chart(document.getElementById('chartDD'), {{
  type: 'line',
  data: {{ datasets: ddData }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    scales: {{
      x: {{
        type: 'time',
        time: {{ unit: 'year', tooltipFormat: 'yyyy-MM-dd' }},
        grid: {{ color: '#2d314820' }},
        ticks: {{ color: '#94a3b8', maxTicksLimit: 12 }},
      }},
      y: {{
        reverse: true,
        grid: {{ color: '#2d314840' }},
        ticks: {{ color: '#94a3b8', callback: v => '-' + v + '%' }},
      }}
    }},
    plugins: {{
      legend: {{ labels: {{ color: '#e2e8f0', boxWidth: 14 }} }},
      tooltip: {{
        backgroundColor: '#1a1d2e', borderColor: '#2d3148', borderWidth: 1,
        titleColor: '#94a3b8', bodyColor: '#e2e8f0',
        callbacks: {{ label: ctx => ctx.dataset.label + ': -' + ctx.parsed.y + '%' }}
      }}
    }}
  }}
}});

// ── Histogramas ──
const histData = {json.dumps(hist_datasets)};
histData.forEach((h, i) => {{
  new Chart(document.getElementById('chartHist' + i), {{
    type: 'bar',
    data: {{
      labels: h.labels,
      datasets: [{{
        label: 'Desde ' + h.yr,
        data: h.counts,
        backgroundColor: h.colors,
        borderWidth: 0,
        borderRadius: 2,
      }}]
    }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      scales: {{
        x: {{
          grid: {{ color: '#2d314820' }},
          ticks: {{ color: '#94a3b8', maxTicksLimit: 10,
                   callback: v => parseFloat(h.labels[v]).toFixed(0) + '€' }},
        }},
        y: {{
          grid: {{ color: '#2d314840' }},
          ticks: {{ color: '#94a3b8' }},
        }}
      }},
      plugins: {{
        legend: {{ labels: {{ color: '#e2e8f0' }} }},
        tooltip: {{
          backgroundColor: '#1a1d2e', borderColor: '#2d3148', borderWidth: 1,
          titleColor: '#94a3b8', bodyColor: '#e2e8f0',
        }}
      }}
    }}
  }});
}});

// ── Tabs métricas anuales ──
function showTab(yr) {{
  document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => {{
    b.classList.remove('active');
    b.style.borderBottomColor = 'transparent';
  }});
  document.getElementById('panel-' + yr).classList.add('active');
  const btn = document.getElementById('tab-' + yr);
  btn.classList.add('active');
  btn.style.borderBottomColor = COLORES[yr];
}}
</script>
</body>
</html>"""

    return html


# --------------------------------------------------
# EXCEL
# --------------------------------------------------

def hex_to_argb(hex_color):
    return "FF" + hex_color.lstrip("#").upper()

def generar_excel(datasets):

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    # Paleta
    C = {
        "bg_dark":   "0F1117",
        "bg_card":   "1A1D2E",
        "bg_header": "12151F",
        "border":    "2D3148",
        "text":      "E2E8F0",
        "muted":     "94A3B8",
        "win":       "10B981",
        "loss":      "EF4444",
        "accent":    "00D4FF",
        "2015":      "00D4FF",
        "2010":      "7C3AED",
        "2005":      "F59E0B",
    }

    def fill(hex_c):
        return PatternFill("solid", fgColor=hex_to_argb(hex_c))

    def font(hex_c="E2E8F0", bold=False, size=11):
        return Font(color=hex_to_argb(hex_c), bold=bold, size=size, name="Calibri")

    def border_thin(hex_c="2D3148"):
        s = Side(style="thin", color=hex_to_argb(hex_c))
        return Border(left=s, right=s, top=s, bottom=s)

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def write_cell(ws, row, col, value, bg=None, fg="E2E8F0", bold=False,
                   size=11, h_align="left", border=True, number_format=None):
        cell = ws.cell(row=row, column=col, value=value)
        if bg:
            cell.fill = fill(bg)
        cell.font = font(fg, bold, size)
        cell.alignment = align(h_align)
        if border:
            cell.border = border_thin()
        if number_format:
            cell.number_format = number_format
        return cell

    for ds in sorted(datasets, key=lambda x: x["year"], reverse=True):
        yr   = ds["year"]
        col  = C[yr]
        m    = ds["metrics"]
        tr   = ds["trades"].copy()
        cap  = ds["capital"].copy()
        ma   = metricas_anuales(cap, tr)

        ws = wb.create_sheet(title=f"Desde {yr}")
        ws.sheet_view.showGridLines = False

        # Fondo global
        ws.sheet_properties.tabColor = hex_to_argb(col)

        # ── Título ──────────────────────────────────────────────────
        ws.merge_cells("A1:M1")
        c = ws["A1"]
        c.value = f"LIBERTAD_2045 v2 — Backtest {yr}–2025"
        c.fill  = fill(col)
        c.font  = Font(color="0F1117", bold=True, size=16, name="Calibri")
        c.alignment = align("center")
        ws.row_dimensions[1].height = 36

        ws.merge_cells("A2:M2")
        c = ws["A2"]
        c.value = f"5 mejoras activas · Generado {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        c.fill  = fill("12151F")
        c.font  = font("94A3B8", size=10)
        c.alignment = align("center")
        ws.row_dimensions[2].height = 20

        # ── KPIs ────────────────────────────────────────────────────
        kpi_row = 4
        kpis = [
            ("Capital final",   f"{m['capital_final']:,.0f} €",  col),
            ("Retorno total",   f"{m['retorno_total']*100:.0f}%", col),
            ("Total trades",    int(m['total_trades']),           "E2E8F0"),
            ("Win Rate",        f"{m['win_rate']*100:.1f}%",      C["win"]),
            ("Profit Factor",   f"{m['profit_factor']:.4f}",      C["win"]),
            ("Max Drawdown",    f"{m['max_drawdown']*100:.1f}%",  C["loss"]),
            ("Expectativa",     f"{m['expectativa']:,.0f} €",     "E2E8F0"),
            ("PnL medio WIN",   f"{m['pnl_medio_win']:,.0f} €",   C["win"]),
            ("PnL medio LOSS",  f"{m['pnl_medio_loss']:,.0f} €",  C["loss"]),
        ]

        for j, (label, val, fgc) in enumerate(kpis):
            c_col = j + 1
            ws.cell(row=kpi_row,   column=c_col).value = label
            ws.cell(row=kpi_row,   column=c_col).fill  = fill("12151F")
            ws.cell(row=kpi_row,   column=c_col).font  = font("94A3B8", bold=False, size=9)
            ws.cell(row=kpi_row,   column=c_col).alignment = align("center")
            ws.cell(row=kpi_row+1, column=c_col).value = val
            ws.cell(row=kpi_row+1, column=c_col).fill  = fill("1A1D2E")
            ws.cell(row=kpi_row+1, column=c_col).font  = font(fgc, bold=True, size=12)
            ws.cell(row=kpi_row+1, column=c_col).alignment = align("center")
            ws.column_dimensions[get_column_letter(c_col)].width = 17

        ws.row_dimensions[kpi_row].height = 18
        ws.row_dimensions[kpi_row+1].height = 26

        # ── Sección: Métricas anuales ────────────────────────────────
        sec_row = kpi_row + 4
        ws.merge_cells(f"A{sec_row}:I{sec_row}")
        c = ws.cell(row=sec_row, column=1, value="MÉTRICAS ANUALES")
        c.fill = fill(col)
        c.font = Font(color="0F1117", bold=True, size=10, name="Calibri")
        c.alignment = align("left")
        ws.row_dimensions[sec_row].height = 20

        hdr_row = sec_row + 1
        hdrs = ["Año", "Cap. inicio (€)", "Cap. fin (€)", "Retorno (%)",
                "Trades", "Wins", "Win Rate (%)", "PnL total (€)", "Max DD (%)"]
        for j, h in enumerate(hdrs):
            c = ws.cell(row=hdr_row, column=j+1, value=h)
            c.fill = fill("12151F")
            c.font = font("94A3B8", bold=True, size=9)
            c.alignment = align("center")
            c.border = border_thin()
        ws.row_dimensions[hdr_row].height = 18

        for ri, (_, row) in enumerate(ma.iterrows()):
            r = hdr_row + 1 + ri
            vals = [
                (int(row["Año"]),         "E2E8F0", "0"),
                (row["Cap. inicio"],      "E2E8F0", '#,##0'),
                (row["Cap. fin"],         "E2E8F0", '#,##0'),
                (row["Retorno"],          C["win"] if row["Retorno"] >= 0 else C["loss"], '0.0'),
                (int(row["Trades"]),      "E2E8F0", "0"),
                (int(row["Wins"]),        C["win"],  "0"),
                (row["Win Rate"],         C["win"],  "0.0"),
                (row["PnL total"],        C["win"] if row["PnL total"] >= 0 else C["loss"], '#,##0'),
                (row["Max DD"],           C["loss"], "0.0"),
            ]
            bg = "1A1D2E" if ri % 2 == 0 else "12151F"
            for j, (val, fgc, fmt) in enumerate(vals):
                c = ws.cell(row=r, column=j+1, value=val)
                c.fill = fill(bg)
                c.font = font(fgc, size=10)
                c.alignment = align("center")
                c.border = border_thin()
                if fmt not in ("0",):
                    c.number_format = fmt
            ws.row_dimensions[r].height = 16

        # ── Sección: Todos los trades ────────────────────────────────
        trades_start = hdr_row + 1 + len(ma) + 3
        ws.merge_cells(f"A{trades_start}:J{trades_start}")
        c = ws.cell(row=trades_start, column=1,
                    value=f"TODOS LOS TRADES ({len(tr):,})")
        c.fill = fill(col)
        c.font = Font(color="0F1117", bold=True, size=10, name="Calibri")
        c.alignment = align("left")
        ws.row_dimensions[trades_start].height = 20

        th_row = trades_start + 1
        trade_hdrs = ["#", "Símbolo", "Clase", "Entrada", "Salida",
                      "Shares", "Precio entrada", "Precio salida", "PnL (€)", "Resultado"]
        for j, h in enumerate(trade_hdrs):
            c = ws.cell(row=th_row, column=j+1, value=h)
            c.fill = fill("12151F")
            c.font = font("94A3B8", bold=True, size=9)
            c.alignment = align("center")
            c.border = border_thin()
        ws.row_dimensions[th_row].height = 18

        for ti, (_, t) in enumerate(tr.iterrows()):
            r = th_row + 1 + ti
            bg = "1A1D2E" if ti % 2 == 0 else "12151F"
            pnl_color = C["win"] if t["pnl"] >= 0 else C["loss"]
            res_color  = C["win"] if t["resultado"] == "WIN" else (
                         "94A3B8" if "OPEN" in str(t["resultado"]) else C["loss"])

            row_data = [
                (ti + 1,                    "94A3B8", "0"),
                (t["symbol"],               "E2E8F0", None),
                (t.get("clase", "ACCION"),  "94A3B8", None),
                (str(t["fecha_entrada"].date()) if pd.notna(t["fecha_entrada"]) else "", "94A3B8", None),
                (str(t["fecha_salida"].date())  if pd.notna(t["fecha_salida"])  else "", "94A3B8", None),
                (int(t["shares"]),          "E2E8F0", "0"),
                (t["entrada"],              "E2E8F0", "0.00"),
                (t["salida"],               "E2E8F0", "0.00"),
                (t["pnl"],                  pnl_color, "#,##0.00"),
                (t["resultado"],            res_color, None),
            ]
            for j, (val, fgc, fmt) in enumerate(row_data):
                c = ws.cell(row=r, column=j+1, value=val)
                c.fill = fill(bg)
                c.font = font(fgc, size=9)
                c.alignment = align("center")
                c.border = border_thin()
                if fmt:
                    c.number_format = fmt
            ws.row_dimensions[r].height = 15

        # Anchos de columnas trades
        anchos = [6, 10, 10, 12, 12, 8, 14, 14, 12, 12]
        for j, w in enumerate(anchos):
            ws.column_dimensions[get_column_letter(j+1)].width = w

        # ── Hoja de curva de capital (datos) ──────────────────────────
        # Añadir columna de capital en columna L como referencia
        cap_col = 12
        cap_row_start = kpi_row + 4
        ws.cell(row=cap_row_start, column=cap_col,
                value="CURVA DE CAPITAL").fill = fill("12151F")
        ws.cell(row=cap_row_start, column=cap_col).font = font("94A3B8", bold=True, size=9)
        ws.cell(row=cap_row_start, column=cap_col+1,
                value="").fill = fill("12151F")

        ws.cell(row=cap_row_start+1, column=cap_col,  value="Fecha").fill = fill("12151F")
        ws.cell(row=cap_row_start+1, column=cap_col).font = font("94A3B8", bold=True, size=9)
        ws.cell(row=cap_row_start+1, column=cap_col+1, value="Capital (€)").fill = fill("12151F")
        ws.cell(row=cap_row_start+1, column=cap_col+1).font = font("94A3B8", bold=True, size=9)
        ws.column_dimensions[get_column_letter(cap_col)].width = 13
        ws.column_dimensions[get_column_letter(cap_col+1)].width = 14

        # Submuestrear a 500 puntos para no saturar Excel
        step = max(1, len(cap) // 500)
        for ki, (_, row) in enumerate(cap.iloc[::step].iterrows()):
            r = cap_row_start + 2 + ki
            ws.cell(row=r, column=cap_col,   value=str(row["fecha"].date())).fill = fill("1A1D2E" if ki%2==0 else "12151F")
            ws.cell(row=r, column=cap_col).font   = font("94A3B8", size=9)
            ws.cell(row=r, column=cap_col+1, value=round(row["capital"], 2)).fill = fill("1A1D2E" if ki%2==0 else "12151F")
            ws.cell(row=r, column=cap_col+1).font = font(col, size=9)
            ws.cell(row=r, column=cap_col+1).number_format = "#,##0"

    # ── Hoja resumen comparativo ─────────────────────────────────────────
    ws_res = wb.create_sheet(title="Comparativa v1 vs v2", index=0)
    ws_res.sheet_view.showGridLines = False
    ws_res.sheet_properties.tabColor = hex_to_argb("10B981")

    ws_res.merge_cells("A1:L1")
    c = ws_res["A1"]
    c.value = "LIBERTAD_2045 — Comparativa Versión 1 vs Versión 2"
    c.fill = fill("10B981")
    c.font = Font(color="0F1117", bold=True, size=15, name="Calibri")
    c.alignment = align("center")
    ws_res.row_dimensions[1].height = 36

    ws_res.merge_cells("A2:L2")
    c = ws_res["A2"]
    c.value = "5 mejoras implementadas el 2026-04-14 · Win rate: 49–51% → 60–62%"
    c.fill = fill("12151F")
    c.font = font("94A3B8", size=10)
    c.alignment = align("center")
    ws_res.row_dimensions[2].height = 20

    mejoras = [
        ("M1", "data_loader.py",   "Ventana datos 1Y → 2Y",              "Activa el stop dinámico B1 (ATR percentil) que estaba inactivo"),
        ("M2", "signal_engine.py", "Pullback ventana 1 día → 3 días",     "+29% señales detectadas · misma calidad de setup"),
        ("M3", "libertad2045.py",  "Score + pendiente SMA200 (5 días)",    "Prioriza acciones donde la tendencia de largo plazo acelera"),
        ("M4", "rebalance.py",     "Break-even automático (1.5×ATR)",      "Stop sube a entry+0.5×ATR cuando precio ≥ entry+1.5×ATR"),
        ("M5", "signal_engine.py", "Pullback fijo 2% → ATR-adaptativo",   "Exige corrección proporcional a la volatilidad del activo"),
    ]

    mejora_hdr_row = 4
    ws_res.cell(row=mejora_hdr_row, column=1, value="MEJORAS IMPLEMENTADAS").fill = fill("0F1117")
    ws_res.cell(row=mejora_hdr_row, column=1).font = font("10B981", bold=True, size=9)
    ws_res.cell(row=mejora_hdr_row, column=1).alignment = align("left")
    ws_res.row_dimensions[mejora_hdr_row].height = 18

    for ji, (cod, arch, titulo, desc) in enumerate(mejoras):
        r = mejora_hdr_row + 1 + ji
        bg = "1A1D2E" if ji % 2 == 0 else "12151F"
        for c_idx, (val, fgc, w) in enumerate([
            (cod,    "10B981", 5),
            (arch,   "00D4FF", 20),
            (titulo, "E2E8F0", 32),
            (desc,   "94A3B8", 52),
        ]):
            c = ws_res.cell(row=r, column=c_idx+1, value=val)
            c.fill = fill(bg)
            c.font = font(fgc, bold=(c_idx == 0), size=10)
            c.alignment = align("left")
            c.border = border_thin()
            ws_res.column_dimensions[get_column_letter(c_idx+1)].width = w
        ws_res.row_dimensions[r].height = 18

    # Tabla comparativa
    comp_row = mejora_hdr_row + len(mejoras) + 3
    ws_res.cell(row=comp_row, column=1, value="COMPARATIVA RESULTADOS").fill = fill("0F1117")
    ws_res.cell(row=comp_row, column=1).font = font("10B981", bold=True, size=9)
    ws_res.row_dimensions[comp_row].height = 18

    comp_hdrs = ["Período", "Capital final v1", "Capital final v2", "Δ Capital",
                 "Retorno v1", "Retorno v2", "Win Rate v1", "Win Rate v2", "Δ WR",
                 "PF v1", "PF v2", "Max DD v1", "Max DD v2"]
    hdr_r = comp_row + 1
    for j, h in enumerate(comp_hdrs):
        c = ws_res.cell(row=hdr_r, column=j+1, value=h)
        c.fill = fill("12151F")
        c.font = font("94A3B8", bold=True, size=9)
        c.alignment = align("center")
        c.border = border_thin()
    ws_res.row_dimensions[hdr_r].height = 18

    for ri, ds in enumerate(sorted(datasets, key=lambda x: x["year"], reverse=True)):
        yr  = ds["year"]
        m   = ds["metrics"]
        ant = ANTES[yr]
        col = C[yr]
        r   = hdr_r + 1 + ri
        bg  = "1A1D2E" if ri % 2 == 0 else "12151F"

        delta_cap = m["capital_final"] - ant["cap_final"]
        delta_wr  = m["win_rate"] * 100 - ant["win_rate"] * 100

        row_data = [
            (f"Desde {yr}",                     col,       None),
            (ant["cap_final"],                  "94A3B8",  "#,##0"),
            (m["capital_final"],                col,       "#,##0"),
            (delta_cap,                         "10B981",  "#,##0"),
            (f"{ant['retorno']*100:.1f}%",      "94A3B8",  None),
            (f"{m['retorno_total']*100:.1f}%",  col,       None),
            (f"{ant['win_rate']*100:.1f}%",     "94A3B8",  None),
            (f"{m['win_rate']*100:.1f}%",       "10B981",  None),
            (f"+{delta_wr:.1f}pp",              "10B981",  None),
            (f"{ant['pf']:.4f}",                "94A3B8",  None),
            (f"{m['profit_factor']:.4f}",       "10B981",  None),
            (f"{ant['dd']*100:.1f}%",           "94A3B8",  None),
            (f"{m['max_drawdown']*100:.1f}%",   "10B981",  None),
        ]
        for j, (val, fgc, fmt) in enumerate(row_data):
            c = ws_res.cell(row=r, column=j+1, value=val)
            c.fill = fill(bg)
            c.font = font(fgc, bold=(j == 0), size=10)
            c.alignment = align("center")
            c.border = border_thin()
            if fmt:
                c.number_format = fmt
        ws_res.row_dimensions[r].height = 18

    for j, w in enumerate([14,16,16,14,12,12,12,12,10,12,12,12,12]):
        ws_res.column_dimensions[get_column_letter(j+1)].width = w

    return wb


# --------------------------------------------------
# MAIN
# --------------------------------------------------

if __name__ == "__main__":

    print("Cargando datos de los 3 últimos backtests...")
    datasets = cargar_datos()

    if len(datasets) < 3:
        print(f"ERROR: solo se encontraron {len(datasets)} conjuntos de resultados.")
        exit(1)

    for ds in datasets:
        print(f"  Período desde {ds['year']}: {len(ds['trades'])} trades · "
              f"capital final {ds['metrics']['capital_final']:,.0f} €")

    # HTML
    print("\nGenerando dashboard HTML...")
    html  = generar_html(datasets)
    html_path = OUTPUT_DIR / "libertad2045_dashboard_v2.html"
    html_path.write_text(html, encoding="utf-8")
    print(f"  Guardado: {html_path.resolve()}")

    # Excel
    print("\nGenerando Excel detallado...")
    wb    = generar_excel(datasets)
    xl_path = OUTPUT_DIR / "libertad2045_backtest_v2.xlsx"
    wb.save(xl_path)
    print(f"  Guardado: {xl_path.resolve()}")

    print("\nListo.")

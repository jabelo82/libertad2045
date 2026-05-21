#!/usr/bin/env python3
"""
Añade la hoja 'Montecarlo' a scripts_libertad.xlsx
con los resultados del análisis Montecarlo (2005-2025).
Preserva el esquema visual oscuro del libro existente.
"""

import os
import glob
import shutil
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# Paleta de colores (extraída del libro existente)
# ─────────────────────────────────────────────────────────────────────────────
C = {
    "bg_header":  "161B22",   # fondo cabeceras de sección
    "bg_row_a":   "1C2128",   # fila alternada A
    "bg_row_b":   "0D1117",   # fila alternada B
    "bg_title":   "0D1117",   # título / subtítulo
    "bg_subhead": "0D2137",   # sub-cabeceras de categoría
    "bg_cell":    "161B22",   # celda destacada
    "fg_blue":    "58A6FF",   # azul cabecera
    "fg_white":   "F0F6FF",   # blanco datos clave
    "fg_grey":    "CDD9E5",   # gris datos secundarios
    "fg_subtle":  "6E7681",   # gris subtítulo
    "fg_green":   "3FB950",   # verde positivo
    "fg_red":     "F85149",   # rojo negativo / alerta
    "fg_amber":   "E3B341",   # ámbar neutral / advertencia
    "fg_purple":  "BC8CFF",   # púrpura (datos MC)
    "fg_teal":    "39D0D8",   # teal (percentiles)
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(hex_color, bold=False, size=11, italic=False):
    return Font(name="Calibri", color=hex_color, bold=bold,
                size=size, italic=italic)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_bottom(color="30363D"):
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def border_full(color="30363D"):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def border_top(color="58A6FF"):
    return Border(top=Side(style="medium", color=color))

# ─────────────────────────────────────────────────────────────────────────────
# Re-ejecutar el Montecarlo para obtener datos frescos
# ─────────────────────────────────────────────────────────────────────────────
RESULTS_DIR = os.path.expanduser("~/PROYECTO_LIBERTAD_2045/backtest_results")
capital_files = sorted(glob.glob(os.path.join(RESULTS_DIR, "expandido_capital_*.csv")))
metrics_files = sorted(glob.glob(os.path.join(RESULTS_DIR, "expandido_metricas_*.csv")))
trades_files  = sorted(glob.glob(os.path.join(RESULTS_DIR, "expandido_trades_*.csv")))

df_cap     = pd.read_csv(capital_files[-1], parse_dates=["fecha"])
df_metrics = pd.read_csv(metrics_files[-1])
df_trades  = pd.read_csv(trades_files[-1])

CAPITAL_INICIAL = float(df_metrics["capital_inicial"].iloc[0])
MAX_DD_REAL     = float(df_metrics["max_drawdown"].iloc[0])
TOTAL_TRADES    = int(df_metrics["total_trades"].iloc[0])
WIN_RATE        = float(df_metrics["win_rate"].iloc[0])
PROFIT_FACTOR   = float(df_metrics["profit_factor"].iloc[0])
EXPECTATIVA     = float(df_metrics["expectativa"].iloc[0])
WIN_N           = int(df_metrics["wins"].iloc[0])
LOSS_N          = int(df_metrics["losses"].iloc[0])

df_cap   = df_cap.sort_values("fecha").reset_index(drop=True)
CAPITAL_REAL = float(df_cap["capital"].iloc[-1])
FECHA_INI    = df_cap["fecha"].iloc[0].strftime("%d/%m/%Y")
FECHA_FIN    = df_cap["fecha"].iloc[-1].strftime("%d/%m/%Y")

equity_real   = df_cap["capital"].values
daily_returns = np.diff(equity_real) / equity_real[:-1]
n_days        = len(daily_returns)

RNG    = np.random.default_rng(seed=42)
N_SIMS = 1000
print(f"Ejecutando {N_SIMS} simulaciones Montecarlo...")

all_shuffles   = RNG.choice(daily_returns, size=(N_SIMS, n_days), replace=True)
final_capitals = np.empty(N_SIMS)
max_drawdowns  = np.empty(N_SIMS)

for i in range(N_SIMS):
    rets   = all_shuffles[i]
    equity = CAPITAL_INICIAL * np.concatenate([[1.0], np.cumprod(1 + rets)])
    final_capitals[i] = equity[-1]
    peak   = np.maximum.accumulate(equity)
    dd     = (equity - peak) / peak
    max_drawdowns[i] = float(dd.min())

dd_abs = np.abs(max_drawdowns)
ruina  = np.mean(final_capitals < CAPITAL_INICIAL)

# Percentiles drawdown
dd_pcts = [10, 25, 50, 75, 90, 95, 99]
dd_vals  = [np.percentile(dd_abs, p) for p in dd_pcts]

# Percentiles capital final
cap_pcts = [5, 10, 25, 50, 75, 90, 95]
cap_vals  = [np.percentile(final_capitals, p) for p in cap_pcts]

# Posición del backtest real en la distribución
pct_rank_cap = np.mean(final_capitals < CAPITAL_REAL)
pct_rank_dd  = np.mean(dd_abs < MAX_DD_REAL)

retorno_real = CAPITAL_REAL / CAPITAL_INICIAL - 1
geo_ret_day  = (CAPITAL_REAL / CAPITAL_INICIAL) ** (1 / n_days) - 1
geo_ret_yr   = (1 + geo_ret_day) ** 252 - 1

print("Montecarlo completado.")
print(f"  DD p50={np.percentile(dd_abs,50):.2%}  p95={np.percentile(dd_abs,95):.2%}  p99={np.percentile(dd_abs,99):.2%}")
print(f"  Cap p50=${np.percentile(final_capitals,50):,.0f}  Real=${CAPITAL_REAL:,.0f}  (percentil {pct_rank_cap:.0%})")
print(f"  Prob. ruina: {ruina:.2%}")

# ─────────────────────────────────────────────────────────────────────────────
# Abrir el libro y crear / reemplazar la hoja
# ─────────────────────────────────────────────────────────────────────────────
XLSX_PATH = os.path.expanduser("~/PROYECTO_LIBERTAD_2045/scripts_libertad.xlsx")
shutil.copy2(XLSX_PATH, XLSX_PATH.replace(".xlsx", "_backup.xlsx"))
print(f"Backup guardado.")

wb = openpyxl.load_workbook(XLSX_PATH)
if "Montecarlo" in wb.sheetnames:
    del wb["Montecarlo"]

ws = wb.create_sheet("Montecarlo")

# ─────────────────────────────────────────────────────────────────────────────
# Helpers de escritura
# ─────────────────────────────────────────────────────────────────────────────
def write(row, col, value, bg, fg, bold=False, h_align="center",
          size=11, italic=False, wrap=False, num_format=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill(bg)
    c.font      = font(fg, bold=bold, size=size, italic=italic)
    c.alignment = align(h_align, "center", wrap=wrap)
    if num_format:
        c.number_format = num_format
    if border:
        c.border = border
    return c

def merge_write(row, col_start, col_end, value, bg, fg,
                bold=False, h_align="center", size=11, italic=False,
                wrap=False, border=None):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = write(row, col_start, value, bg, fg,
              bold=bold, h_align=h_align, size=size,
              italic=italic, wrap=wrap, border=border)
    return c

def blank_row(row, cols, bg):
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = fill(bg)

# ─────────────────────────────────────────────────────────────────────────────
# Anchos de columna
# ─────────────────────────────────────────────────────────────────────────────
col_widths = {
    1: 32,   # Métrica / label
    2: 22,   # Valor / backtest real
    3: 22,   # MC mediana / p50
    4: 22,   # MC p95 / p5
    5: 22,   # MC p99 / p95
    6: 24,   # Nota / descripción
}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Alto de filas por defecto
ws.sheet_format.defaultRowHeight = 18

# ─────────────────────────────────────────────────────────────────────────────
# FILA 1: Título principal
# ─────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 28
merge_write(1, 1, 6,
            "LIBERTAD 2045 — Análisis de Montecarlo (2005–2025)",
            C["bg_header"], C["fg_blue"], bold=True, size=14)

# FILA 2: Subtítulo
ws.row_dimensions[2].height = 18
merge_write(2, 1, 6,
            f"1.000 simulaciones · Bootstrap con reposición sobre {n_days:,} retornos diarios · "
            f"Generado: {datetime.today().strftime('%d/%m/%Y')}",
            C["bg_title"], C["fg_subtle"], italic=True, size=10)

# FILA 3: separador
blank_row(3, 6, C["bg_title"])
ws.row_dimensions[3].height = 6

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN A: Parámetros del backtest
# ─────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[4].height = 20
merge_write(4, 1, 6, "▸  PARÁMETROS DEL BACKTEST",
            C["bg_subhead"], C["fg_amber"], bold=True, size=11)

# Cabeceras
row = 5
ws.row_dimensions[row].height = 18
for c, txt in enumerate(["Parámetro", "Valor"], start=1):
    write(row, c, txt, C["bg_header"], C["fg_blue"], bold=True)
for c in range(3, 7):
    ws.cell(row=row, column=c).fill = fill(C["bg_header"])

backtest_params = [
    ("Período",              f"{FECHA_INI} → {FECHA_FIN}",   C["fg_white"]),
    ("Capital inicial",      f"${CAPITAL_INICIAL:,.0f}",      C["fg_grey"]),
    ("Capital final (real)", f"${CAPITAL_REAL:,.0f}",         C["fg_green"]),
    ("Retorno total",        f"{retorno_real:+.0%}",          C["fg_green"]),
    ("Retorno anualizado",   f"{geo_ret_yr:.1%}",             C["fg_green"]),
    ("Drawdown máximo real", f"{MAX_DD_REAL:.2%}",            C["fg_red"]),
    ("Total trades",         f"{TOTAL_TRADES:,}",             C["fg_grey"]),
    ("Wins / Losses",        f"{WIN_N:,} / {LOSS_N:,}",      C["fg_grey"]),
    ("Win rate",             f"{WIN_RATE:.1%}",               C["fg_amber"]),
    ("Profit factor",        f"{PROFIT_FACTOR:.4f}",         C["fg_amber"]),
    ("Expectativa / trade",  f"${EXPECTATIVA:,.2f}",          C["fg_amber"]),
    ("Días de trading",      f"{n_days:,}",                   C["fg_grey"]),
]

for i, (param, val, color) in enumerate(backtest_params):
    r = row + 1 + i
    ws.row_dimensions[r].height = 18
    bg = C["bg_row_a"] if i % 2 == 0 else C["bg_row_b"]
    write(r, 1, param, bg, C["fg_grey"],  h_align="left")
    write(r, 2, val,   bg, color, bold=(color == C["fg_green"]))
    for c in range(3, 7):
        ws.cell(row=r, column=c).fill = fill(bg)

row = row + 1 + len(backtest_params)

# Separador
blank_row(row, 6, C["bg_title"]); ws.row_dimensions[row].height = 8
row += 1

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN B: Resumen de riesgo — tabla comparativa principal
# ─────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[row].height = 20
merge_write(row, 1, 6, "▸  RESUMEN DE RIESGO — BACKTEST REAL vs. MONTECARLO",
            C["bg_subhead"], C["fg_amber"], bold=True, size=11)
row += 1

# Cabeceras
ws.row_dimensions[row].height = 20
headers_b = ["Métrica", "Backtest Real", "MC Mediana (p50)",
             "MC Worst 5% (p5/p95)", "MC Worst 1% (p1/p99)", "Descripción"]
for c, h in enumerate(headers_b, start=1):
    write(row, c, h, C["bg_header"], C["fg_blue"], bold=True,
          size=11, border=border_bottom())
row += 1

p1_cap  = np.percentile(final_capitals, 1)
p99_cap = np.percentile(final_capitals, 99)

risk_rows = [
    # (label, real, mc_p50, mc_worst5, mc_worst1, desc)
    (
        "Capital final",
        f"${CAPITAL_REAL:,.0f}",
        f"${np.percentile(final_capitals,50):,.0f}",
        f"${np.percentile(final_capitals,5):,.0f}",
        f"${p1_cap:,.0f}",
        "Capital al cierre del período",
    ),
    (
        "Retorno total",
        f"{retorno_real:+.0%}",
        f"{np.percentile(final_capitals,50)/CAPITAL_INICIAL-1:+.0%}",
        f"{np.percentile(final_capitals,5)/CAPITAL_INICIAL-1:+.0%}",
        f"{p1_cap/CAPITAL_INICIAL-1:+.0%}",
        "Respecto al capital inicial",
    ),
    (
        "Drawdown máximo",
        f"{MAX_DD_REAL:.2%}",
        f"{np.percentile(dd_abs,50):.2%}",
        f"{np.percentile(dd_abs,95):.2%}",
        f"{np.percentile(dd_abs,99):.2%}",
        "Caída pico-valle sobre equity",
    ),
    (
        "Prob. de ruina",
        "—",
        f"{ruina:.2%}",
        f"{ruina:.2%}",
        f"{ruina:.2%}",
        "Capital < capital inicial",
    ),
    (
        "Percentil real (capital)",
        f"{pct_rank_cap:.0%}",
        "—",
        "—",
        "—",
        "Posición del backtest real en MC",
    ),
    (
        "Percentil real (drawdown)",
        f"{pct_rank_dd:.0%}",
        "—",
        "—",
        "—",
        "Posición del DD real en MC",
    ),
    (
        "Capital en riesgo total",
        f"${CAPITAL_REAL - np.percentile(final_capitals,5):,.0f}",
        "—",
        "—",
        "—",
        "Real vs. p5 MC (downside 5%)",
    ),
]

for i, (label, real, p50v, p5v, p1v, desc) in enumerate(risk_rows):
    ws.row_dimensions[row].height = 18
    bg = C["bg_row_a"] if i % 2 == 0 else C["bg_row_b"]
    write(row, 1, label, bg, C["fg_white"],  bold=True, h_align="left")
    write(row, 2, real,  bg, C["fg_amber"])
    write(row, 3, p50v,  bg, C["fg_green"])
    write(row, 4, p5v,   bg, C["fg_red"])
    write(row, 5, p1v,   bg, C["fg_red"])
    write(row, 6, desc,  bg, C["fg_subtle"], italic=True, h_align="left", size=10)
    row += 1

blank_row(row, 6, C["bg_title"]); ws.row_dimensions[row].height = 8
row += 1

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN C: Distribución de drawdowns — tabla de percentiles
# ─────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[row].height = 20
merge_write(row, 1, 6, "▸  DISTRIBUCIÓN DE DRAWDOWNS MÁXIMOS (1.000 simulaciones)",
            C["bg_subhead"], C["fg_amber"], bold=True, size=11)
row += 1

ws.row_dimensions[row].height = 20
for c, h in enumerate(["Percentil", "Drawdown máximo", "Simulaciones con DD ≤ valor",
                        "Interpretación", "", ""], start=1):
    write(row, c, h if c <= 4 else "",
          C["bg_header"], C["fg_blue"], bold=True, border=border_bottom())
row += 1

dd_interp = {
    10: "9 de cada 10 escenarios mejores",
    25: "3 de cada 4 escenarios mejores",
    50: "Escenario típico / mediano",
    75: "1 de cada 4 escenarios peor",
    90: "1 de cada 10 escenarios peor",
    95: "Escenario adverso (1 en 20)",
    99: "Escenario extremo (1 en 100)",
}

for i, (p, v) in enumerate(zip(dd_pcts, dd_vals)):
    ws.row_dimensions[row].height = 18
    bg = C["bg_row_a"] if i % 2 == 0 else C["bg_row_b"]
    is_real_zone = (p == 99)   # el DD real está en percentil ~98

    label_color = C["fg_red"] if p >= 95 else (C["fg_amber"] if p >= 50 else C["fg_teal"])
    write(row, 1, f"p{p}",   bg, label_color, bold=True)
    write(row, 2, f"{v:.2%}", bg, label_color, bold=(p >= 95))
    write(row, 3, f"{p}%",   bg, C["fg_grey"])
    write(row, 4, dd_interp.get(p, ""), bg, C["fg_subtle"], italic=True,
          h_align="left", size=10)
    for c in range(5, 7):
        ws.cell(row=row, column=c).fill = fill(bg)
    row += 1

# Fila especial: DD real del backtest
ws.row_dimensions[row].height = 18
bg = C["bg_cell"]
merge_write(row, 1, 1, "REAL (backtest)", bg, C["fg_amber"], bold=True, h_align="left")
write(row, 2, f"{MAX_DD_REAL:.2%}", bg, C["fg_amber"], bold=True)
write(row, 3, f"≈ {pct_rank_dd:.0%}", bg, C["fg_amber"])
write(row, 4, f"El DD real supera al {pct_rank_dd:.0%} de simulaciones",
      bg, C["fg_amber"], italic=True, h_align="left", size=10)
for c in range(5, 7):
    ws.cell(row=row, column=c).fill = fill(bg)
row += 1

blank_row(row, 6, C["bg_title"]); ws.row_dimensions[row].height = 8
row += 1

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN D: Distribución de capitales finales
# ─────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[row].height = 20
merge_write(row, 1, 6, "▸  DISTRIBUCIÓN DE CAPITALES FINALES (1.000 simulaciones)",
            C["bg_subhead"], C["fg_amber"], bold=True, size=11)
row += 1

ws.row_dimensions[row].height = 20
for c, h in enumerate(["Percentil", "Capital final", "Retorno total",
                        "Múltiplo sobre inicio", "Interpretación", ""], start=1):
    write(row, c, h if c <= 5 else "",
          C["bg_header"], C["fg_blue"], bold=True, border=border_bottom())
row += 1

cap_interp = {
    5:  "Peor 5%: escenario muy adverso",
    10: "Peor 10%: escenario adverso",
    25: "Cuartil inferior",
    50: "Escenario mediano",
    75: "Cuartil superior",
    90: "Mejor 10%: escenario favorable",
    95: "Mejor 5%: escenario muy favorable",
}

for i, (p, v) in enumerate(zip(cap_pcts, cap_vals)):
    ws.row_dimensions[row].height = 18
    bg   = C["bg_row_a"] if i % 2 == 0 else C["bg_row_b"]
    ret  = v / CAPITAL_INICIAL - 1
    mult = v / CAPITAL_INICIAL

    if p <= 10:
        color = C["fg_red"]
    elif p <= 25:
        color = C["fg_amber"]
    elif p == 50:
        color = C["fg_white"]
    else:
        color = C["fg_green"]

    write(row, 1, f"p{p}",         bg, color, bold=(p in (5, 95)))
    write(row, 2, f"${v:,.0f}",    bg, color, bold=(p in (5, 95)))
    write(row, 3, f"{ret:+.0%}",   bg, color)
    write(row, 4, f"{mult:.0f}×",  bg, C["fg_grey"])
    write(row, 5, cap_interp.get(p, ""), bg, C["fg_subtle"],
          italic=True, h_align="left", size=10)
    ws.cell(row=row, column=6).fill = fill(bg)
    row += 1

# Fila especial: capital real del backtest
ws.row_dimensions[row].height = 18
bg = C["bg_cell"]
merge_write(row, 1, 1, "REAL (backtest)", bg, C["fg_amber"], bold=True, h_align="left")
write(row, 2, f"${CAPITAL_REAL:,.0f}", bg, C["fg_amber"], bold=True)
write(row, 3, f"{retorno_real:+.0%}", bg, C["fg_amber"])
write(row, 4, f"{CAPITAL_REAL/CAPITAL_INICIAL:.0f}×", bg, C["fg_amber"])
write(row, 5, f"Percentil {pct_rank_cap:.0%} en la distribución MC",
      bg, C["fg_amber"], italic=True, h_align="left", size=10)
ws.cell(row=row, column=6).fill = fill(bg)
row += 1

blank_row(row, 6, C["bg_title"]); ws.row_dimensions[row].height = 8
row += 1

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN E: Tabla de estadísticas descriptivas (distribuciones)
# ─────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[row].height = 20
merge_write(row, 1, 6, "▸  ESTADÍSTICAS DESCRIPTIVAS DE LAS DISTRIBUCIONES",
            C["bg_subhead"], C["fg_amber"], bold=True, size=11)
row += 1

ws.row_dimensions[row].height = 20
for c, h in enumerate(["Estadístico", "Drawdown máximo",
                        "Capital final", "Retorno total", "", ""], start=1):
    write(row, c, h if c <= 4 else "",
          C["bg_header"], C["fg_blue"], bold=True, border=border_bottom())
row += 1

stats = [
    ("Mínimo",             f"{dd_abs.min():.2%}",             f"${final_capitals.min():,.0f}",        f"{final_capitals.min()/CAPITAL_INICIAL-1:+.0%}"),
    ("Media",              f"{dd_abs.mean():.2%}",            f"${final_capitals.mean():,.0f}",       f"{final_capitals.mean()/CAPITAL_INICIAL-1:+.0%}"),
    ("Mediana (p50)",      f"{np.median(dd_abs):.2%}",        f"${np.median(final_capitals):,.0f}",   f"{np.median(final_capitals)/CAPITAL_INICIAL-1:+.0%}"),
    ("Máximo",             f"{dd_abs.max():.2%}",             f"${final_capitals.max():,.0f}",        f"{final_capitals.max()/CAPITAL_INICIAL-1:+.0%}"),
    ("Desv. estándar",     f"{dd_abs.std():.2%}",             f"${final_capitals.std():,.0f}",        "—"),
    ("Asimetría",          f"{float(pd.Series(dd_abs).skew()):.2f}",  f"{float(pd.Series(final_capitals).skew()):.2f}", "—"),
    ("Prob. ruina (< ini)", f"—",                             f"{ruina:.2%}",                         "—"),
    ("N simulaciones",     f"{N_SIMS:,}",                     f"{N_SIMS:,}",                          "—"),
]

for i, (stat, dd_v, cap_v, ret_v) in enumerate(stats):
    ws.row_dimensions[row].height = 18
    bg = C["bg_row_a"] if i % 2 == 0 else C["bg_row_b"]
    write(row, 1, stat,  bg, C["fg_grey"],  h_align="left")
    write(row, 2, dd_v,  bg, C["fg_teal"])
    write(row, 3, cap_v, bg, C["fg_purple"])
    write(row, 4, ret_v, bg, C["fg_purple"])
    for c in range(5, 7):
        ws.cell(row=row, column=c).fill = fill(bg)
    row += 1

blank_row(row, 6, C["bg_title"]); ws.row_dimensions[row].height = 8
row += 1

# ─────────────────────────────────────────────────────────────────────────────
# SECCIÓN F: Nota metodológica
# ─────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[row].height = 20
merge_write(row, 1, 6, "▸  NOTA METODOLÓGICA",
            C["bg_subhead"], C["fg_amber"], bold=True, size=11)
row += 1

notas = [
    ("Método",    "Bootstrap con reposición (resample) sobre retornos diarios de la curva de capital.",     C["fg_grey"]),
    ("Por qué",   "La permutación pura da siempre el mismo capital final (producto es conmutativo). "
                  "El bootstrap con reposición genera verdadera variación al repetir/omitir días.",         C["fg_subtle"]),
    ("Supuesto",  "Se asume que la distribución histórica de retornos diarios es representativa del futuro.", C["fg_subtle"]),
    ("Limitación","No captura cambios de régimen ni correlaciones temporales a largo plazo.",               C["fg_subtle"]),
    ("Semilla RNG","numpy.default_rng(seed=42) — resultados reproducibles.",                               C["fg_grey"]),
    ("Archivo",   os.path.basename(capital_files[-1]),                                                     C["fg_grey"]),
]

for i, (key, val, color) in enumerate(notas):
    ws.row_dimensions[row].height = 30
    bg = C["bg_row_a"] if i % 2 == 0 else C["bg_row_b"]
    write(row, 1, key, bg, C["fg_blue"], bold=True, h_align="left")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws.cell(row=row, column=2, value=val)
    c.fill      = fill(bg)
    c.font      = font(color, size=10, italic=(color == C["fg_subtle"]))
    c.alignment = align("left", "center", wrap=True)
    row += 1

# Padding final
for r in range(row, row + 3):
    blank_row(r, 6, C["bg_title"])

# ─────────────────────────────────────────────────────────────────────────────
# Freeze panes, tab color, zoom
# ─────────────────────────────────────────────────────────────────────────────
ws.freeze_panes = "A4"
ws.sheet_properties.tabColor = "58A6FF"
ws.sheet_view.zoomScale = 90

# ─────────────────────────────────────────────────────────────────────────────
# Guardar
# ─────────────────────────────────────────────────────────────────────────────
wb.save(XLSX_PATH)
print(f"\n✓ Hoja 'Montecarlo' añadida a: {XLSX_PATH}")
print(f"  Filas escritas: {row}")
print(f"  Hojas del libro: {wb.sheetnames}")
